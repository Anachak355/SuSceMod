# -*- coding: utf-8 -*-
"""
@author: Anasua Chakraborty
"""

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import rcParams
import numpy as np
import seaborn as sns
import openpyxl

rcParams['font.family'] = 'Arial'

# Function to extract significance markers from the 'Sig' column
def get_significance(sig_str):

    # Strip leading/trailing whitespace
    sig_str = sig_str.strip()

    # Map the significance to positions
    significance_map = []
    
    for i, char in enumerate(sig_str):
        if char == '*':
            significance_map.append('†')  # Mark with special symbol
        else:
            significance_map.append('')  # No symbol for non-significant
    
    return significance_map

# Function to plot the coefficients with significance
def plot_coefficients_simplified(coefficients, feature_names, legend_labels, significant_flags, palette=["#ffa183", "#ca2525", "#514747"], title='MLR Coefficients for Expansion (Cls0-1)', xlabel='Coefficient Value', save_path=None):

    sns.set_palette(palette)

    # Plotting
    fig, ax = plt.subplots(figsize=(11, 10))
    bar_width = 2 # Width of each bar

    # Ensure the number of coefficients matches the number of feature names
    for i, label in enumerate(legend_labels):
        gap = 1.5
        y_positions = np.arange(len(feature_names)) * (len(legend_labels) * bar_width + gap) + (i * bar_width)

        # Plot each coefficient for the class
        bars = ax.barh(y_positions, coefficients[:, i], height=bar_width, label=label, alpha=0.9)
        flags = [row[i] for row in significant_flags]
        
        # Add significance markers directly on bars
        for j, (bar, significant) in enumerate(zip(bars, flags)):
            if significant:
                bar.set_edgecolor('black')
                bar.set_hatch('//')

    # Set y-axis ticks and labels
    y_ticks = np.arange(len(feature_names)) * (len(legend_labels) * bar_width + gap) + (len(legend_labels) * bar_width / 2)
    
    # Set the y-ticks and labels
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(feature_names, fontsize=26)
        
    # Set x-axis labels
    ax.set_xlabel(xlabel, fontsize=26)
    ax.tick_params(axis='x', labelsize=26)
    
    # Add legend outside the plot
    legend = plt.legend(title=None, loc='lower right', fontsize=23)
    plt.setp(legend.get_title(), fontsize=24)  # Set legend title font size

    # Set title
    # ax.set_title(title, fontsize=22)

    # Add grid
    ax.grid(axis='x', linestyle='--', alpha=0.5)

    # Remove spines
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    
    plt.xlim([-2,4])
    
    # Show the plot
    plt.tight_layout()
    if save_path:
        plt.savefig(save_path, dpi=600, format='pdf')  # Save as PDF
    plt.show()

# Function to read the named tables from an Excel workbook
def process_excel_tables(file_path):
    # Load the workbook with openpyxl
    wb = openpyxl.load_workbook(file_path)

    # Get the table names (we assume these are defined as named ranges)
    table_names = wb.sheetnames  # Updated for sheetnames instead of get_sheet_names()

    # Initialize list for coefficients and significance markers
    all_coefficients = []
    all_significant_flags = []
    feature_names = []

    # Loop through each table and extract the data
    for table_name in table_names:
        if table_name in wb.sheetnames:  # Ensure the sheet exists
            sheet = wb[table_name]
            
            # Extract the range of the named table
            table = sheet.tables[table_name]
            table_range = table.ref  # Get the range (e.g., 'A1:D12')

            # Extract data using the table range, we will iterate through the range of rows and columns
            data = []
            for row in sheet[table_range]:
                row_data = [cell.value for cell in row]  # Extract cell values
                data.append(row_data)

            # Convert to a DataFrame
            df = pd.DataFrame(data[1:], columns=data[0])  # Skip first row for header

            # Extract feature names (Variable column)
            feature_names = df['Variable'].values.tolist()

            # Extract coefficients (cls-1, cls-2, cls-3, etc.) - skip the first two columns
            coefficients = df.iloc[:, 2:-1].values  # Extract coefficients (skip the first two columns)

            # Extract significance column (Sig)
            sig_column = df['Sig'].values.tolist()

            # Convert significance to special symbols
            significant_flags = [get_significance(sig) for sig in sig_column]

            all_coefficients.append(coefficients)
            all_significant_flags.append(significant_flags)

    return all_coefficients, all_significant_flags, feature_names, table_names

###########################################

file_path = "...xlsx" # Please input here
all_coefficients, all_significant_flags, feature_names, table_names = process_excel_tables(file_path)

# Plot first three tables
for i, table_name in enumerate(table_names[:3]):
    plot_coefficients_simplified(
        coefficients=all_coefficients[i],
        feature_names=feature_names,
        legend_labels=[f'Class {i} - Class {i+ j+1}' for j in range(all_coefficients[i].shape[1])],
        # palette=["#e38a8a", "#b22234", "#6b6b6b"],
        palette=["#a8d5ba", "#fbb4ae", "#c8b9e6"],
        significant_flags=all_significant_flags[i],
        title=f'MNL Coefficients transitioning from Class {i}',  # Title for each table
        xlabel='Coefficient Value',
        save_path=f'Class{i}_coefficients_plot.pdf'  # Save as PDF
    )
