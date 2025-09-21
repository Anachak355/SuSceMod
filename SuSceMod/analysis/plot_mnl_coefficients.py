import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import rcParams
import numpy as np
import seaborn as sns
import openpyxl

rcParams['font.family'] = 'Arial'

def get_significance(sig_str):
    sig_str = sig_str.strip()
    significance_map = []
    
    for i, char in enumerate(sig_str):
        if char == '*':
            significance_map.append('†')
        else:
            significance_map.append('')
    return significance_map

def plot_coefficients_simplified(
    coefficients, 
    feature_names, 
    legend_labels, 
    significant_flags, 
    palette=["#ffa183", "#ca2525", "#514747"], 
    title='MLR Coefficients for Expansion (Cls0-1)', 
    xlabel='Coefficient Value', 
    save_path=None):

    sns.set_palette(palette)

    fig, ax = plt.subplots(figsize=(11, 10))
    bar_width = 2

    for i, label in enumerate(legend_labels):
        gap = 1.5
        y_positions = np.arange(len(feature_names)) * (len(legend_labels) * bar_width + gap) + (i * bar_width)

        bars = ax.barh(y_positions, coefficients[:, i], height=bar_width, label=label, alpha=0.9)
        flags = [row[i] for row in significant_flags]
        
        for j, (bar, significant) in enumerate(zip(bars, flags)):
            if significant:
                bar.set_edgecolor('black')
                bar.set_hatch('//')

    y_ticks = np.arange(len(feature_names)) * (len(legend_labels) * bar_width + gap) + (len(legend_labels) * bar_width / 2)
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(feature_names, fontsize=26)
        
    ax.set_xlabel(xlabel, fontsize=26)
    ax.tick_params(axis='x', labelsize=26)
    
    legend = plt.legend(title=None, loc='lower right', fontsize=23)
    plt.setp(legend.get_title(), fontsize=24)  # Set legend title font size

    ax.set_title(title, fontsize=22)
    ax.grid(axis='x', linestyle='--', alpha=0.5)

    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    
    plt.tight_layout()
    if save_path:
        plt.savefig(save_path, dpi=600, format='pdf')
    plt.show()

def process_excel_tables(file_path):
    wb = openpyxl.load_workbook(file_path)

    table_names = wb.sheetnames

    all_coefficients = []
    all_significant_flags = []
    feature_names = []

    for table_name in table_names:
        if table_name in wb.sheetnames:
            sheet = wb[table_name]
            
            table = sheet.tables[table_name]
            table_range = table.ref

            data = []
            for row in sheet[table_range]:
                row_data = [cell.value for cell in row]
                data.append(row_data)

            df = pd.DataFrame(data[1:], columns=data[0])

            feature_names = df['Variable'].values.tolist()

            coefficients = df.iloc[:, 2:-1].values

            sig_column = df['Sig'].values.tolist()

            significant_flags = [get_significance(sig) for sig in sig_column]

            all_coefficients.append(coefficients)
            all_significant_flags.append(significant_flags)

    return all_coefficients, all_significant_flags, feature_names, table_names

###########################################

def plot_mnl_coefficients(file_path):
    all_coefficients, all_significant_flags, feature_names, table_names = process_excel_tables(file_path)
    
    # Plot first three tables
    for i, table_name in enumerate(table_names[:3]):
        plot_coefficients_simplified(
            coefficients=all_coefficients[i],
            feature_names=feature_names,
            legend_labels=[f'Class {i} - Class {i+ j+1}' for j in range(all_coefficients[i].shape[1])],
            # palette=["#e38a8a", "#b22234", "#6b6b6b"],
            palette=["#a8d5ba", "#fbb4ae", "#c8b9e6"],
            significant_flags=all_significant_flags[i],
            title=f'MNL Coefficients transitioning from Class {i}',  # Title for each table
            xlabel='Coefficient Value',
            save_path=f'Class{i}_coefficients_plot.pdf'  # Save as PDF
        )
